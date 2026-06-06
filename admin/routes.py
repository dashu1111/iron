from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file, abort
from flask_login import login_required, current_user
from models import db, User, Overtime, Leave, Bonus
from datetime import datetime
from utils import role_required, export_excel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from collections import defaultdict
from io import BytesIO

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/dashboard')
@login_required
@role_required('admin')
def dashboard():
    total_users = User.query.count()
    total_ot = Overtime.query.count()
    total_leave = Leave.query.count()
    return render_template('admin/dashboard.html', total_users=total_users, total_ot=total_ot, total_leave=total_leave)

@admin_bp.route('/users')
@login_required
@role_required('admin')
def user_manage():
    users = User.query.all()
    return render_template('admin/user_manage.html', users=users)

@admin_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def user_add():
    if request.method == 'POST':
        user = User(
            username=request.form['username'],
            name=request.form['name'],
            role=request.form['role'],
            group=request.form.get('group'),
            department=request.form.get('department')
        )
        user.set_password(request.form['password'])
        db.session.add(user)
        db.session.commit()
        flash('账号已创建', 'success')
        return redirect(url_for('admin.user_manage'))
    return render_template('admin/user_add.html')

@admin_bp.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def user_edit(id):
    user = db.session.get(User, id)
    if not user:
        abort(404)
    if request.method == 'POST':
        user.username = request.form['username']
        user.name = request.form['name']
        user.role = request.form['role']
        user.group = request.form.get('group')
        user.department = request.form.get('department')
        db.session.commit()
        flash('账号已更新', 'success')
        return redirect(url_for('admin.user_manage'))
    return render_template('admin/user_edit.html', user=user)

@admin_bp.route('/users/<int:id>/reset_password')
@login_required
@role_required('admin')
def reset_password(id):
    user = db.session.get(User, id)
    if not user:
        abort(404)
    user.set_password('123456')  # 默认密码
    db.session.commit()
    flash(f'{user.name} 的密码已重置为 123456', 'success')
    return redirect(url_for('admin.user_manage'))

@admin_bp.route('/users/<int:id>/toggle')
@login_required
@role_required('admin')
def toggle_user(id):
    user = db.session.get(User, id)
    if not user:
        abort(404)
    user.enabled = not user.enabled
    db.session.commit()
    flash('账号状态已切换', 'success')
    return redirect(url_for('admin.user_manage'))

@admin_bp.route('/bonus/assign', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def bonus_assign():
    users = User.query.filter(User.role != 'admin').all()
    if request.method == 'POST':
        year_month = request.form['year_month']
        for u in users:
            amount = request.form.get(f'amount_{u.id}', 0)
            reason = request.form.get(f'reason_{u.id}', '')
            if float(amount) > 0:
                bonus = Bonus(
                    user_id=u.id,
                    year_month=year_month,
                    amount=float(amount),
                    reason=reason,
                    assigned_by=current_user.id
                )
                db.session.add(bonus)
        db.session.commit()
        flash('奖金分配成功', 'success')
        return redirect(url_for('admin.dashboard'))
    return render_template('admin/bonus_assign.html', users=users)

@admin_bp.route('/all_records')
@login_required
@role_required('admin')
def all_records():
    overtimes = Overtime.query.order_by(Overtime.date.desc()).limit(100).all()
    leaves = Leave.query.order_by(Leave.start_date.desc()).limit(100).all()
    bonuses = Bonus.query.order_by(Bonus.year_month.desc()).limit(100).all()
    return render_template('admin/all_records.html', overtimes=overtimes, leaves=leaves, bonuses=bonuses)

@admin_bp.route('/export', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def export_data():
    if request.method == 'POST':
        data_type = request.form['data_type']
        # start_month / end_month / department 可留待扩展，此处仅示例全部导出
        if data_type == 'overtime':
            records = Overtime.query.filter_by(status='approved').order_by(Overtime.date).all()
            user_records = defaultdict(lambda: defaultdict(list))
            for rec in records:
                key = f"{rec.ot_type}-{rec.shift}"
                user_records[rec.user_id][key].append(rec.date)

            users = {u.id: u for u in User.query.filter(User.id.in_(user_records.keys())).all()}

            all_columns = set()
            for ud in user_records.values():
                all_columns.update(ud.keys())
            type_order = ['在线内包', '在线外包', '行车', '离线', '支撑检修']
            sorted_columns = []
            for t in type_order:
                for s in ['白班', '夜班']:
                    key = f"{t}-{s}"
                    if key in all_columns:
                        sorted_columns.append(key)
                        all_columns.remove(key)
            sorted_columns.extend(sorted(all_columns))

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "加班统计汇总"
            headers = ['姓名', '作业区', '班组'] + sorted_columns
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for user_id, type_dict in user_records.items():
                user = users.get(user_id)
                if not user:
                    continue
                row = [user.name, user.department or '', user.group or '']
                for col_key in sorted_columns:
                    dates = type_dict.get(col_key, [])
                    date_str = ', '.join(f"{d.month}.{d.day}" for d in sorted(dates)) if dates else ''
                    row.append(date_str)
                ws.append(row)

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            for i in range(4, 4 + len(sorted_columns)):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, as_attachment=True, download_name='加班汇总表.xlsx')
        flash('导出功能请根据需求完善', 'info')
        return redirect(url_for('admin.export_data'))
    return render_template('admin/export.html')